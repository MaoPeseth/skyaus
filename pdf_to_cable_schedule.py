"""
PDF Cable Schedule Converter
Converts PDF cable schedules to the 2-column Excel format.
Usage:
  python pdf_to_cable_schedule.py path/to/file.pdf
  python pdf_to_cable_schedule.py path/to/folder/
"""

import sys, os, re, json
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    os.system('pip install pdfplumber -q')
    import pdfplumber

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system('pip install openpyxl -q')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color='FF0D2240', end_color='FF0D2240', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Calibri', size=10)
WARN_FONT = Font(name='Calibri', size=10, color='CC0000')

def find_cable_headers(row):
    cn_col = cl_col = -1
    for ci, cell in enumerate(row):
        if cell is None: continue
        hl = str(cell).strip().lower()
        if hl in ('cable number', 'cable no.', 'cable no', 'cable id', 'cable #'):
            cn_col = ci
        if hl == 'cable label' or hl.startswith('cable label') or ('cable label' in hl):
            cl_col = ci
    return cn_col, cl_col

def is_valid_cable_number(val):
    if not val: return False
    s = str(val).strip()
    if not s: return False
    if s.lower() in ('cable number', 'cable no', 'cable id'): return False
    if re.match(r'^[A-Z]{1,6}-\d+', s, re.I): return True
    return bool(re.match(r'^[A-Z]{2,6}\s*\d+', s, re.I))

def is_section_header(val):
    if not val: return False
    s = str(val).strip().upper()
    return s in ('AC POWER', 'DC POWER', 'COAXIAL', 'FIBER OPTIC', 'FIBRE OPTIC', 'ANTENNA / COAXIAL NETWORK', 'WIRELESS NETWORK')

def has_cable_label_separator(val):
    if not val: return False
    s = str(val).strip()
    s = re.sub(r'\s+', ' ', s)
    return ' - ' in s

def normalize_label(val):
    if not val: return ''
    s = str(val).strip()
    s = re.sub(r'\s+', ' ', s)
    return s

def extract_from_pdf(pdf_path):
    results = []
    errors = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            if not tables:
                errors.append(f"  Page {page_num}: no tables found")
                continue
            for table in tables:
                if not table or len(table) < 2: continue
                header_row = None
                cn_col = cl_col = -1
                for ri, row in enumerate(table):
                    if not row: continue
                    c, l = find_cable_headers(row)
                    if c >= 0:
                        header_row = ri
                        cn_col = c
                        cl_col = l if l >= 0 else cl_col
                        break
                if header_row is None: continue
                last_cn = ''
                for ri in range(header_row + 1, len(table)):
                    row = table[ri]
                    if not row: continue
                    raw_cn = str(row[cn_col]).strip() if cn_col < len(row) and row[cn_col] is not None else ''
                    raw_cl = str(row[cl_col]).strip() if cl_col >= 0 and cl_col < len(row) and row[cl_col] is not None else ''
                    cn = raw_cn.strip()
                    cl = normalize_label(raw_cl)
                    if not cl: continue  # skip completely empty rows
                    # Section headers like "AC POWER", "DC POWER" — skip
                    if is_section_header(cl) or is_section_header(cn):
                        last_cn = ''
                        continue
                    # If cable number is empty, carry forward the last one (multi-core cables)
                    if not cn:
                        if last_cn:
                            cn = last_cn
                        else:
                            continue  # no previous cable to attach to
                    else:
                        if not is_valid_cable_number(cn):
                            last_cn = ''
                            continue
                        last_cn = cn
                    if not has_cable_label_separator(cl):
                        errors.append(f"  Page {page_num}, row {ri+1}: '{cn}' label missing ' - ' separator: '{cl}'")
                    results.append((cn, cl))
    return results, errors

def write_excel(results, errors, pdf_path, output_dir):
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(output_dir, f"{base}_cable_schedule.xlsx")
    # Remove old file if it exists (might be locked from previous run)
    for attempt in range(3):
        try:
            if os.path.exists(out_path): os.remove(out_path)
            break
        except PermissionError:
            import time; time.sleep(0.5)
    log_path = os.path.join(output_dir, f"{base}_review.log")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cable Schedule'

    headers = ['cable number', 'cable label']
    col_widths = [22, 60]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[chr(64+ci)].width = w

    for ri, (cn, cl) in enumerate(results, 2):
        for ci, val in enumerate([cn, cl], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.auto_filter.ref = f"A1:B{len(results)+1}"
    ws.freeze_panes = 'A2'

    wb.save(out_path)

    with open(log_path, 'w') as f:
        f.write(f"PDF: {pdf_path}\n")
        f.write(f"Converted: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"Valid rows extracted: {len(results)}\n")
        f.write(f"Issues found: {len(errors)}\n\n")
        if errors:
            f.write("=== ISSUES ===\n")
            for e in errors:
                f.write(f"{e}\n")
        if results:
            prefixes = {}
            for cn, _ in results:
                p = re.match(r'^([A-Z]+)', cn, re.I)
                if p:
                    key = p.group(1).upper()
                    prefixes[key] = prefixes.get(key, 0) + 1
            f.write(f"\n=== CABLE SUMMARY ===\n")
            for p, c in sorted(prefixes.items()):
                f.write(f"  {p}: {c}\n")
            f.write(f"  TOTAL: {len(results)}\n")

    return out_path, log_path

def process_pdf(pdf_path, output_dir):
    print(f"\n{'='*60}")
    print(f"Processing: {pdf_path}")
    print(f"{'='*60}")
    results, errors = extract_from_pdf(pdf_path)
    if not results:
        print("  No valid cable rows found.")
        return
    out_path, log_path = write_excel(results, errors, pdf_path, output_dir)
    prefixes = {}
    for cn, _ in results:
        p = re.match(r'^([A-Z]+)', cn, re.I)
        if p:
            key = p.group(1).upper()
            prefixes[key] = prefixes.get(key, 0) + 1
    print(f"  Extracted: {len(results)} cables")
    for p, c in sorted(prefixes.items()):
        print(f"    {p}: {c}")
    print(f"  Issues: {len(errors)}")
    print(f"  Output: {out_path}")
    print(f"  Log:    {log_path}")

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = os.path.join(os.path.dirname(os.path.abspath(input_path if os.path.isfile(input_path) else input_path)), 'converted')
    os.makedirs(output_dir, exist_ok=True)

    if os.path.isfile(input_path):
        if not input_path.lower().endswith('.pdf'):
            print("Error: Input must be a PDF file.")
            sys.exit(1)
        process_pdf(input_path, output_dir)
    elif os.path.isdir(input_path):
        pdfs = [f for f in os.listdir(input_path) if f.lower().endswith('.pdf')]
        if not pdfs:
            print(f"No PDF files found in {input_path}")
            sys.exit(1)
        print(f"Found {len(pdfs)} PDF(s) in {input_path}")
        for pdf in sorted(pdfs):
            process_pdf(os.path.join(input_path, pdf), output_dir)
    else:
        print(f"Error: {input_path} not found.")
        sys.exit(1)

    print(f"\nAll outputs saved to: {output_dir}")
    print("Done.")

if __name__ == '__main__':
    main()
